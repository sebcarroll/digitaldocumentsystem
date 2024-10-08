"""Module for extracting text from various file formats stored on Google Drive."""

import os
import io
import docx2txt
from typing import Union, List
from io import BytesIO
import csv
import xlrd
import openpyxl
import chardet

from app.services.google_drive.core import DriveCore
from googleapiclient.http import MediaIoBaseDownload
from langchain_community.document_loaders import (
    Docx2txtLoader,
    CSVLoader,
    TextLoader,
    UnstructuredExcelLoader,
    PyPDFLoader
)
from PyPDF2 import PdfReader
from langchain_community.document_loaders import PyPDFLoader
from langchain.schema import Document

class FileExtractor:
    """
    A class to extract text from various file formats including .docx, .doc, PDFs,
    Google Docs, Google Sheets, Excel files, and files stored on Google Drive.
    """

    def __init__(self, drive_core: DriveCore):
        """
        Initialize FileExtractor with a DriveCore instance.

        Args:
            drive_core (DriveCore): An instance of the DriveCore class for Google Drive operations.
        """
        self.drive_core = drive_core

    def convert_google_doc_to_docx(self, file_id: str) -> BytesIO:
        """
        Convert a Google Doc to .docx format.

        Args:
            file_id (str): The ID of the Google Doc file.

        Returns:
            BytesIO: A file-like object containing the .docx content.

        Raises:
            Exception: If there's an error during the conversion process.
        """
        request = self.drive_core.drive_service.files().export_media(
            fileId=file_id,
            mimeType='application/vnd.openxmlformats-officedocument.wordprocessingml.document'
        )
        file = BytesIO(request.execute())
        return file

    def convert_google_sheet_to_xlsx(self, file_id: str) -> BytesIO:
        """
        Convert a Google Sheet to .xlsx format.

        Args:
            file_id (str): The ID of the Google Sheet file.

        Returns:
            BytesIO: A file-like object containing the .xlsx content.

        Raises:
            Exception: If there's an error during the conversion process.
        """
        request = self.drive_core.drive_service.files().export_media(
            fileId=file_id,
            mimeType='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        file = io.BytesIO(request.execute())
        return file

    def load_document(self, file: Union[str, BytesIO], file_type: str) -> List[Document]:
        """
        Load a document using the appropriate Langchain loader.

        Args:
            file (Union[str, BytesIO]): A file path, URL, or BytesIO object containing the document content.
            file_type (str): The type of the file (e.g., 'docx', 'csv', 'txt', 'pdf', 'xlsx').

        Returns:
            List[Document]: A list of Langchain Document objects.

        Raises:
            ValueError: If the file type is not supported.
            Exception: If there's an error during the loading process.
        """
        if file_type == 'docx':
            if isinstance(file, BytesIO):
                text = docx2txt.process(file)
                return [Document(page_content=text)]
            else:
                loader = Docx2txtLoader(file)
        elif file_type == 'csv':
            if isinstance(file, BytesIO):
                detected = chardet.detect(file.getvalue())
                encoding = detected['encoding']
                csv_reader = csv.reader(io.StringIO(file.getvalue().decode(encoding)))
                rows = list(csv_reader)
                text = "\n".join([",".join(row) for row in rows])
                return [Document(page_content=text)]
            else:
                loader = CSVLoader(file)
        elif file_type == 'txt':
            if isinstance(file, BytesIO):
                text = file.getvalue().decode('utf-8')
                return [Document(page_content=text)]
            else:
                loader = TextLoader(file)
        elif file_type == 'pdf':
            if isinstance(file, BytesIO):
                pdf_reader = PdfReader(file)
                documents = []
                for page in pdf_reader.pages:
                    text = page.extract_text()
                    documents.append(Document(page_content=text))
                return documents
            else:
                loader = PyPDFLoader(file)
        elif file_type in ['xlsx', 'xls']:
            if isinstance(file, BytesIO):
                if file_type == 'xlsx':
                    workbook = openpyxl.load_workbook(file)
                    sheets = workbook.sheetnames
                else:  # xls
                    workbook = xlrd.open_workbook(file_contents=file.getvalue())
                    sheets = workbook.sheet_names()
                
                text = []
                for sheet_name in sheets:
                    if file_type == 'xlsx':
                        sheet = workbook[sheet_name]
                        sheet_text = "\n".join([" ".join(str(cell.value) for cell in row) for row in sheet.iter_rows()])
                    else:  # xls
                        sheet = workbook.sheet_by_name(sheet_name)
                        sheet_text = "\n".join([" ".join(str(sheet.cell_value(row, col)) for col in range(sheet.ncols)) for row in range(sheet.nrows)])
                    text.append(f"Sheet: {sheet_name}\n{sheet_text}")
                
                return [Document(page_content="\n\n".join(text))]
            else:
                loader = UnstructuredExcelLoader(file)
        else:
            raise ValueError(f"Unsupported file type: {file_type}")
        
        return loader.load()

    def download_file_from_google_drive(self, file_id: str, file_name: str) -> str:
        """
        Download a file from Google Drive.

        Args:
            file_id (str): The ID of the file in Google Drive.
            file_name (str): The name to save the file as locally.

        Returns:
            str: The path to the downloaded file.

        Raises:
            Exception: If there's an error during the download process.
        """
        request = self.drive_core.drive_service.files().get_media(fileId=file_id)
        fh = io.FileIO(file_name, 'wb')
        downloader = MediaIoBaseDownload(fh, request)
        
        done = False
        while not done:
            status, done = downloader.next_chunk()
        
        fh.close()
        return file_name

    def extract_text_from_drive_file(self, file_id: str, file_name: str) -> str:
        """
        Download and extract text from a Google Drive file using Langchain loaders.

        This method determines the file type, downloads or converts the file as necessary,
        and then extracts the text content using appropriate loaders.

        Args:
            file_id (str): The ID of the file in Google Drive.
            file_name (str): The name of the file.

        Returns:
            str: The extracted text content from the file. Returns an empty string if
                extraction fails or no text is found.

        Raises:
            Exception: If there's an error during the extraction process that cannot be handled.
        """
        # Get file metadata to determine the MIME type
        file_metadata = self.drive_core.drive_service.files().get(fileId=file_id, fields='mimeType').execute()
        mime_type = file_metadata['mimeType']

        if mime_type == 'application/vnd.google-apps.document':
            file = self.convert_google_doc_to_docx(file_id)
            file_extension = 'docx'
        elif mime_type == 'application/vnd.google-apps.spreadsheet':
            file = self.convert_google_sheet_to_xlsx(file_id)
            file_extension = 'xlsx'
        elif mime_type in ['application/pdf', 'text/csv', 'text/plain', 
            'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            'application/vnd.ms-excel',
            'application/vnd.openxmlformats-officedocument.wordprocessingml.document']:
            request = self.drive_core.drive_service.files().get_media(fileId=file_id)
            file = io.BytesIO(request.execute())
            if mime_type == 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet':
                file_extension = 'xlsx'
            elif mime_type == 'application/vnd.ms-excel':
                file_extension = 'xls'
            elif mime_type == 'application/vnd.openxmlformats-officedocument.wordprocessingml.document':
                file_extension = 'docx'
            else:
                file_extension = mime_type.split('/')[-1]
        else:
            raise ValueError(f"Unsupported MIME type: {mime_type}")

        documents = self.load_document(file, file_extension)
        extracted_text = "\n\n".join([doc.page_content for doc in documents])
        
        return extracted_text